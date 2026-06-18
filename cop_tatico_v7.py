# =============================================================================
# COP TÁTICO — Common Operational Picture  v7.0
# Exercício de Simulação Construtiva
#
# Novidades v7.0:
#   - Lê diretamente do Google Sheets (sem exportar nada)
#   - IDs começando com "_EX-" são ignorados (exemplos de referência)
#   - ID gerado automaticamente via fórmula no Google Sheets
#   - Fallback automático para arquivo local se offline
#   - Planilha: https://docs.google.com/spreadsheets/d/1HirteDnnhZ2YU1z-aqnTmTj-wUlgKM77m3zjdRz3f3M
#   - Tudo mais igual à v5 (UTM 22S, xlsx local, hqtf, mobilidade etc.)
# =============================================================================

import os
import io
import hashlib
import urllib.request
import urllib.error
import urllib.parse

from qgis.core import (
    QgsVectorFileWriter,
    QgsLayerTree, QgsLayerTreeGroup,
    QgsProject, QgsVectorLayer, QgsField, QgsFeature,
    QgsGeometry, QgsPointXY, QgsFields,
    QgsSingleSymbolRenderer, QgsSymbol,
    QgsSvgMarkerSymbolLayer, QgsProperty,
    QgsPalLayerSettings, QgsVectorLayerSimpleLabeling,
    QgsTextFormat, QgsTextBufferSettings,
    QgsCoordinateReferenceSystem, QgsCoordinateTransform,
)
from PyQt5.QtCore import QVariant, QTimer
from PyQt5.QtGui import QColor, QFont

# =============================================================================
# CONFIGURAÇÃO
# =============================================================================

# ID da planilha Google Sheets
SHEETS_ID    = '1HirteDnnhZ2YU1z-aqnTmTj-wUlgKM77m3zjdRz3f3M'
NOME_ABA     = 'Lançamento'

# URL montada após imports para garantir que urllib.parse está disponível
# (necessário para codificar acentos no nome da aba, ex: Lançamento → La%C3%A7amento)
def _montar_url_sheets():
    aba_encoded = urllib.parse.quote(NOME_ABA)
    return (
        f'https://docs.google.com/spreadsheets/d/{SHEETS_ID}'
        f'/export?format=xlsx&sheet={aba_encoded}'
    )

# Fallback local — usado se não houver internet
CAMINHO_LOCAL = r'C:\exercicio\situacao_tatica.xlsx'

NOME_CAMADA           = 'COP Tático'
CAMINHO_GPKG          = r'C:\exercicio\cop_tatico.gpkg'  # arquivo único com todas as camadas
PORTA_SERVIDOR        = 2525
TAMANHO_SIMBOLO       = 24
INTERVALO_ATUALIZACAO = 30      # segundos
CRS_ENTRADA           = 'EPSG:4326'
CRS_UTM               = 'EPSG:31982'  # UTM fuso 22S

# Estrutura da planilha
LINHA_CABECALHO    = 3
LINHA_DADOS_INICIO = 4
CAMPOS_OBRIGATORIOS = ['id', 'hostilidade', 'natureza']

# Coluna B da planilha — timestamp_edicao gerado pelo Apps Script
# O script lê mas não valida — é preenchimento automático
PREFIXO_SEPARADOR  = '▼'

# =============================================================================
# TRANSFORMAÇÃO UTM 22S → WGS84
# =============================================================================

_crs_utm   = QgsCoordinateReferenceSystem(CRS_UTM)
_crs_wgs84 = QgsCoordinateReferenceSystem(CRS_ENTRADA)
_transform = QgsCoordinateTransform(
    _crs_utm, _crs_wgs84, QgsProject.instance()
)

def utm_para_latlon(e, n):
    p = _transform.transform(QgsPointXY(e, n))
    return p.x(), p.y()  # lon, lat


def _parse_gms(texto):
    """
    Extrai grau decimal de qualquer variante GMS/GMD.
    Suporta: 26°14'25.5"S | 26 14 25.5S | S 26 14 25.5 | 26°14.5'S
    Retorna (decimal, hemisferio) ou (None, None).
    """
    import re
    t = str(texto).strip().upper()
    hem = None
    for h in ['S', 'N', 'W', 'E', 'O']:
        if t.startswith(h):
            hem = h; t = t[1:].strip(); break
        if t.endswith(h):
            hem = h; t = t[:-1].strip(); break
    nums = re.findall(r'\d+(?:[.,]\d+)?', t)
    if not nums:
        return None, None
    graus = float(nums[0].replace(',', '.'))
    mins  = float(nums[1].replace(',', '.')) if len(nums) > 1 else 0.0
    segs  = float(nums[2].replace(',', '.')) if len(nums) > 2 else 0.0
    return graus + mins / 60.0 + segs / 3600.0, hem


def _aplicar_hemisferio(valor, hem, tipo):
    """Aplica sinal negativo conforme hemisfério. Assume Sul/Oeste se ausente."""
    if tipo == 'lat':
        if hem == 'S' or (hem is None and valor > 0):
            return -abs(valor)
        return abs(valor)
    else:
        if hem in ('W', 'O') or (hem is None and valor > 0):
            return -abs(valor)
        return abs(valor)


def _detectar_formato(e_raw, n_raw):
    """
    Detecta o formato das coordenadas e retorna (lon, lat, fmt).

    Formatos suportados nos campos utm_e / utm_n:
      1. UTM 22S        665378 / 7191882
      2. Grau Decimal   -25.4284 / -49.2731  (com ou sem S/N/W/E)
      3. GMS            26°14'25.5"S / 50°20'53.8"W  (qualquer variante)
      4. GMD            25°25.7'S / 49°16.6'W
    """
    import re
    e = str(e_raw).strip()
    n = str(n_raw).strip()

    # Remove separadores de milhar e normaliza vírgula decimal
    def so_numero(v):
        return re.sub(r'[^\d.,-]', '', v).replace(',', '.')

    e_num = so_numero(e)
    n_num = so_numero(n)

    # 1. UTM 22S — ambos numéricos, N > 1.000.000
    try:
        ef = float(e_num)
        nf = float(n_num)
        if 100_000 <= ef <= 900_000 and 1_000_000 <= nf <= 9_999_999:
            lon, lat = utm_para_latlon(ef, nf)
            return lon, lat, 'UTM 22S'
    except (ValueError, TypeError):
        pass

    # 2. Grau Decimal puro
    try:
        ef = float(e_num)
        nf = float(n_num)
        # utm_e=lat, utm_n=lon (ordem mais comum em GPSs brasileiros)
        if -90 <= ef <= 90 and -180 <= nf <= 180:
            return nf, ef, 'Grau Decimal'
        # utm_e=lon, utm_n=lat
        if -180 <= ef <= 180 and -90 <= nf <= 90:
            return ef, nf, 'Grau Decimal'
    except (ValueError, TypeError):
        pass

    # 3. GMS / GMD
    lat_dec, hem_lat = _parse_gms(e)
    lon_dec, hem_lon = _parse_gms(n)
    if lat_dec is not None and lon_dec is not None:
        lat_f = _aplicar_hemisferio(lat_dec, hem_lat, 'lat')
        lon_f = _aplicar_hemisferio(lon_dec, hem_lon, 'lon')
        if -90 <= lat_f <= 90 and -180 <= lon_f <= 180:
            return lon_f, lat_f, 'GMS/GMD'

    return None, None, None


def parse_coordenadas(linha):
    """
    Lê utm_e e utm_n e retorna (lon, lat) em WGS84.
    Detecta automaticamente: UTM 22S, Grau Decimal, GMS ou GMD.
    Fallback para colunas lon/lat de versões anteriores.
    """
    e_raw = linha.get('utm_e', '').strip()
    n_raw = linha.get('utm_n', '').strip()

    if e_raw and n_raw:
        lon, lat, fmt = _detectar_formato(e_raw, n_raw)
        if lon is not None:
            return lon, lat
        raise ValueError(
            f'Formato não reconhecido — utm_e="{e_raw}" utm_n="{n_raw}". '
            f'Use UTM (665378 / 7191882), '
            f'Grau Decimal (-25.43 / -49.27) ou '
            f'GMS (26°14\'25"S / 50°20\'53"W)'
        )

    # Fallback: colunas lon/lat (compatibilidade com versões anteriores)
    lon_raw = linha.get('lon', '').strip()
    lat_raw = linha.get('lat', '').strip()
    if lon_raw and lat_raw:
        try:
            return float(lon_raw.replace(',', '.')), float(lat_raw.replace(',', '.'))
        except ValueError as ex:
            raise ValueError(f'Lat/lon inválido: {ex}')

    raise ValueError('Sem coordenadas (utm_e/utm_n ou lon/lat)')

# =============================================================================
# TABELAS DE CONVERSÃO
# =============================================================================

HOSTILIDADE = {
    'AMIGO':'03','HOSTIL':'06','NEUTRO':'04',
    'DESCONHECIDO':'01','SUSPEITO':'05','PRESUMIDO':'02',
}
DIMENSAO_CAMPO = {
    'UNIDADE':'10','EQUIPAMENTO':'15','INSTALACAO':'20',
    'AEREO':'05','INDIVIDUO':'27',
}
SITUACAO = {
    'CONFIRMADA':'0','ESTIMADA':'1','PLANEJADA':'1','SUSPEITA':'1',
}
ESCALAO = {
    'NONE':'00','SQD':'11','SEC':'12','PEL':'13',
    'CIA':'14','BIA':'14','ESC':'14','BN':'15',
    'GRU':'16','BDA':'17','DIV':'18','CRP':'19','EX':'20','':'00',
}
HQTF = {
    '':'0','Nenhum':'0','Simulacro / Feint':'1',
    'HQ / Posto de Comando':'2','HQ + Simulacro':'3',
    'Força-Tarefa':'4','Força-Tarefa + Simulacro':'5',
    'Força-Tarefa + HQ':'6','Força-Tarefa + HQ + Simul':'7',
}
NATUREZA_MAP = {
    'Infantaria':('10','121100'),'Infantaria Paraquedista':('10','121101'),
    'Infantaria Motorizada':('10','121102'),'Infantaria Mecanizada':('10','121103'),
    'Infantaria Anfíbia':('10','121104'),'Infantaria de Selva / Rangers':('10','121105'),
    'Forças Especiais':('10','121800'),'Comandos':('10','121700'),
    'Blindado (Armored)':('10','121200'),'Blindado Anfíbio':('10','121201'),
    'Cavalaria':('10','121300'),'Cavalaria Blindada':('10','121301'),
    'Cavalaria Aérea':('10','121302'),'Reconhecimento':('10','121900'),
    'Artilharia de Campanha':('10','130300'),'Artilharia Autopropulsada':('10','130301'),
    'Artilharia Rebocada':('10','130302'),'Artilharia de Foguetes':('10','130700'),
    'Artilharia Antiaérea (AAAe)':('10','130100'),'Art. Antiaérea Autopropulsada':('10','130101'),
    'Morteiros':('10','130800'),'Engenharia':('10','140700'),
    'Engenharia de Combate (Pesada)':('10','140701'),'Engenharia de Construção':('10','140702'),
    'Engenharia Topográfica':('10','140703'),'Comunicações / Sinal':('10','111000'),
    'Comunicações Rádio':('10','111001'),'Guerra Eletrônica (GE)':('10','150500'),
    'Guerra Eletrônica — Ataque':('10','150501'),'Guerra Eletrônica — Suporte':('10','150502'),
    'Inteligência':('10','150300'),'Vigilância':('10','150301'),
    'Defesa QBRN':('10','140100'),'Defesa QBRN — Descontaminação':('10','140101'),
    'Suprimento':('10','163600'),'Manutenção':('10','163700'),
    'Saúde / Médico':('10','163800'),'Transporte (Unidade)':('10','163900'),
    'Posto de Comando (genérico)':('10','120000'),'PC — Corpo de Exército':('10','120100'),
    'PC — Divisão':('10','120200'),'PC — Brigada':('10','120300'),
    'PC — Batalhão':('10','120400'),'C2 / Comando e Controle':('10','110000'),
    'Ligação':('10','110500'),'Tropa Especial':('10','111400'),
    'Carro de Combate (MBT)':('15','120105'),'Viatura Blindada de Combate (VBC)':('15','120101'),
    'Transporte Blindado de Pessoal (TBP/APC)':('15','120103'),'Helicóptero':('15','250000'),
    'Ponte (equipamento)':('15','130100'),'Retroescavadeira / Motoniveladora':('15','130800'),
    'Equipamento Limpa-Minas':('15','130900'),'Mina Terrestre':('15','210100'),
    'IED / Artefato Explosivo':('15','210400'),'Radar (equipamento)':('15','220300'),
    'Viatura de Utilidade':('15','140100'),'Caminhão Semi-Reboque':('15','140600'),
    'Obus':('15','110900'),'Lançador de Mísseis':('15','111000'),
    'Morteiro (equipamento)':('15','111400'),'Lançador de Foguetes Múltiplos':('15','111600'),
    'Instalação Militar (genérica)':('20','110000'),'Depósito de Munição':('20','110300'),
    'C3I / Centro de Controle':('20','110500'),'Instalação QBRN':('20','110600'),
    'Ponto de Coleta de Emergência':('20','111100'),'Instalação de Radar':('20','111700'),
    'Base Militar':('20','120801'),'Aeródromo / Aeroporto':('20','120802'),
    'Hospital':('20','120701'),'Posto Médico':('20','120702'),
    'Telecomunicações':('20','121200'),'Torre de Transmissão':('20','121202'),
    'Porto / Harbor':('20','121307'),'Ferrovia / Estação':('20','121308'),
    'Túnel':('20','121313'),'Represa':('20','110900'),
    'Instalação de Combustível':('20','111600'),
    'Desconhecido / Não identificado':('10','000000'),
}

# =============================================================================
# FUNÇÕES AUXILIARES
# =============================================================================

def resolve_natureza(nat_raw):
    nat = str(nat_raw).strip()
    if nat in NATUREZA_MAP:
        ss, ec = NATUREZA_MAP[nat]
        return ss, ec.ljust(10,'0')[:10]
    digits = nat.replace('/','').replace('-','')
    if digits.isdigit():
        if len(digits)==6:  return '10', digits.ljust(10,'0')
        if len(digits)==10: return '10', digits
        if len(digits)==20: return '10', digits[10:20]
    print(f'  ⚠ Natureza não reconhecida: "{nat}" — usando Desconhecido')
    return '10', '0000000000'

def monta_sidc(hostilidade, dimensao, situacao, escalao,
               natureza_raw, hqtf_nome=''):
    ss, entity_10d = resolve_natureza(natureza_raw)
    h  = HOSTILIDADE.get(str(hostilidade).strip().upper(), '01')
    s  = SITUACAO.get(str(situacao).strip().upper(), '0')
    e  = ESCALAO.get(str(escalao).strip().upper(), '00')
    hq = HQTF.get(str(hqtf_nome).strip(), '0')
    dim = DIMENSAO_CAMPO.get(str(dimensao).strip().upper(), '10')
    if dim not in ('05','27'): dim = ss
    return f'10{h}{dim}{s}{hq}{e}{entity_10d}'

def _str(val):
    if val is None: return ''
    return str(val).strip()

# =============================================================================
# DOWNLOAD DO GOOGLE SHEETS
# =============================================================================

def baixar_sheets(url, timeout=10):
    """
    Baixa a planilha do Google Sheets como bytes xlsx.
    Retorna bytes ou None se falhar.
    """
    req = urllib.request.Request(
        url,
        headers={'User-Agent': 'Mozilla/5.0 (COP-Tatico/6.0)'}
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            dados = resp.read()
            print(f'  ↓ Google Sheets: {len(dados)//1024} KB baixados')
            return dados
    except urllib.error.URLError as e:
        print(f'  ⚠ Sem acesso ao Google Sheets: {e.reason}')
        return None
    except Exception as e:
        print(f'  ⚠ Erro ao baixar planilha: {e}')
        return None

# =============================================================================
# LEITURA XLSX — de bytes (Sheets) ou de arquivo local
# =============================================================================

def _ler_xlsx_bytes(dados_bytes):
    """Lê xlsx a partir de bytes em memória (sem salvar em disco)."""
    from openpyxl import load_workbook
    wb = load_workbook(
        filename=io.BytesIO(dados_bytes),
        read_only=True,
        data_only=True
    )
    return _extrair_registros(wb)

def _ler_xlsx_arquivo(caminho):
    """Lê xlsx de arquivo local."""
    from openpyxl import load_workbook
    wb = load_workbook(caminho, read_only=True, data_only=True)
    return _extrair_registros(wb)

def _extrair_registros(wb):
    """
    Extrai registros da aba NOME_ABA.
    Ignora separadores, linhas vazias e exemplos sem dados reais.
    """
    if NOME_ABA not in wb.sheetnames:
        abas = wb.sheetnames
        raise ValueError(
            f'Aba "{NOME_ABA}" não encontrada. '
            f'Abas disponíveis: {abas}'
        )

    ws = wb[NOME_ABA]
    linhas = list(ws.iter_rows(values_only=True))

    # Cabeçalhos da linha 3
    cabecalho_row = linhas[LINHA_CABECALHO - 1]
    colunas = [_str(c).lower() for c in cabecalho_row]

    registros = []
    vazias = 0

    for row in linhas[LINHA_DADOS_INICIO - 1:]:
        linha = {
            colunas[i]: _str(v)
            for i, v in enumerate(row)
            if i < len(colunas)
        }

        # Para após 5 linhas vazias consecutivas
        if all(v == '' for v in linha.values()):
            vazias += 1
            if vazias >= 5: break
            continue
        vazias = 0

        # Ignora separadores visuais (▼)
        if any(_str(v).startswith(PREFIXO_SEPARADOR)
               for v in linha.values()):
            continue

        # Ignora exemplos de referência — IDs começando com '_EX-'
        # Estes ficam na planilha como guia para os operadores
        # mas nunca aparecem no mapa
        id_val = linha.get('id', '')
        if id_val.upper().startswith('_EX-') or id_val.upper().startswith('_EX_'):
            continue

        # Ignora linhas sem campos obrigatórios
        if all(linha.get(c,'') == '' for c in CAMPOS_OBRIGATORIOS):
            continue

        # Ignora entidades marcadas como inativas (Segue Ativo = NÃO)
        # Vazio ou SIM = aparece no mapa; NÃO = some do mapa (histórico mantido)
        segue_ativo = linha.get('segue_ativo', '').strip().upper()
        if segue_ativo == 'NÃO' or segue_ativo == 'NAO':
            continue

        registros.append(linha)

    wb.close()
    return registros

# =============================================================================
# FONTE DE DADOS — Google Sheets com fallback local
# =============================================================================

def ler_planilha_online(url=None, fallback=CAMINHO_LOCAL):
    """
    Tenta Google Sheets primeiro.
    Se falhar e existir arquivo local, usa o local.
    Retorna (registros, fonte) onde fonte é 'online' ou 'local'.
    """
    if url is None:
        url = _montar_url_sheets()
    dados = baixar_sheets(url)
    if dados:
        try:
            registros = _ler_xlsx_bytes(dados)
            return registros, 'online'
        except Exception as e:
            print(f'  ⚠ Erro ao processar Sheets online: {e}')

    # Fallback para arquivo local
    if os.path.exists(fallback):
        print(f'  → Usando arquivo local: {os.path.basename(fallback)}')
        registros = _ler_xlsx_arquivo(fallback)
        return registros, 'local'

    raise RuntimeError(
        'Sem acesso ao Google Sheets e arquivo local não encontrado.\n'
        f'  Verifique a conexão ou salve a planilha em: {fallback}'
    )

# =============================================================================
# CAMPOS DA CAMADA QGIS
# =============================================================================

def _definir_campos():
    campos = QgsFields()
    for nome, tipo in [
        ('id',QVariant.String),('designacao',QVariant.String),
        ('sidc',QVariant.String),('hostilidade',QVariant.String),
        ('dimensao',QVariant.String),('natureza',QVariant.String),
        ('escalao',QVariant.String),('hqtf',QVariant.String),
        ('escalao_sup',QVariant.String),('situacao',QVariant.String),
        ('mobilidade',QVariant.String),('cond_operac',QVariant.String),
        ('utm_e',QVariant.Double),('utm_n',QVariant.Double),
        ('segue_ativo',QVariant.String),('timestamp_edicao',QVariant.String),('obs',QVariant.String),('fonte',QVariant.String),
    ]:
        campos.append(QgsField(nome, tipo))
    return campos

# =============================================================================
# CONVERSÃO: registros → feições QGIS
# =============================================================================

def _registros_para_feicoes(registros, campos_camada):
    feicoes = []
    for i, linha in enumerate(registros, start=1):
        try:
            lon, lat = parse_coordenadas(linha)
        except ValueError as e:
            print(f'  ⚠ Registro {i} ({linha.get("id","?")}): {e}')
            continue

        try:
            utm_e = float(linha.get('utm_e','0').replace(',','.') or 0)
            utm_n = float(linha.get('utm_n','0').replace(',','.') or 0)
        except ValueError:
            utm_e = utm_n = 0.0

        sidc = monta_sidc(
            hostilidade  = linha.get('hostilidade','DESCONHECIDO'),
            dimensao     = linha.get('dimensao','UNIDADE'),
            situacao     = linha.get('situacao','CONFIRMADA'),
            escalao      = linha.get('escalao',''),
            natureza_raw = linha.get('natureza','Desconhecido / Não identificado'),
            hqtf_nome    = linha.get('hqtf',''),
        )

        feat = QgsFeature(campos_camada)
        feat.setGeometry(QgsGeometry.fromPointXY(QgsPointXY(lon, lat)))
        feat.setAttributes([
            linha.get('id',''), linha.get('designacao',''), sidc,
            linha.get('hostilidade',''), linha.get('dimensao',''),
            linha.get('natureza',''), linha.get('escalao',''),
            linha.get('hqtf',''), linha.get('escalao_sup',''),
            linha.get('situacao',''), linha.get('mobilidade',''),
            linha.get('cond_operac',''), utm_e, utm_n,
            linha.get('segue_ativo',''),
            linha.get('timestamp_edicao',''),
            linha.get('obs',''), linha.get('fonte',''),
        ])
        feicoes.append(feat)
    return feicoes

# =============================================================================
# SIMBOLOGIA
# =============================================================================

def _aplicar_simbologia(camada):
    expr = (
        f"'http://localhost:{PORTA_SERVIDOR}/symbol/' || \"sidc\" || '.svg'"
        f" || '?designation=' || \"designacao\""
        f" || '&higher_formation=' || \"escalao_sup\""
        f" || CASE WHEN \"cond_operac\" != '' "
        f"   THEN '&condition=' || \"cond_operac\" ELSE '' END"
        f" || CASE WHEN \"mobilidade\" != '' "
        f"   THEN '&mobility=' || \"mobilidade\" ELSE '' END"
    )
    svg_lyr = QgsSvgMarkerSymbolLayer('')
    svg_lyr.setSize(TAMANHO_SIMBOLO)
    svg_lyr.setDataDefinedProperty(
        QgsSvgMarkerSymbolLayer.PropertyName,
        QgsProperty.fromExpression(expr)
    )
    sym = QgsSymbol.defaultSymbol(camada.geometryType())
    sym.changeSymbolLayer(0, svg_lyr)
    camada.setRenderer(QgsSingleSymbolRenderer(sym))

# =============================================================================
# ETIQUETAS — QGIS 3.44
# =============================================================================

def _aplicar_etiquetas(camada):
    cfg = QgsPalLayerSettings()
    cfg.fieldName = (
        "CASE WHEN \"escalao_sup\" != '' "
        "THEN \"designacao\" || ' / ' || \"escalao_sup\" "
        "ELSE \"designacao\" END"
    )
    cfg.isExpression = True
    cfg.placement = QgsPalLayerSettings.Placement.AroundPoint
    cfg.yOffset = 4.0
    fmt = QgsTextFormat()
    fmt.setFont(QFont('Arial', 8)); fmt.setSize(8)
    fmt.setColor(QColor('#1a1a1a'))
    buf = QgsTextBufferSettings()
    buf.setEnabled(True); buf.setSize(1.0); buf.setColor(QColor('white'))
    fmt.setBuffer(buf); cfg.setFormat(fmt)
    camada.setLabeling(QgsVectorLayerSimpleLabeling(cfg))
    camada.setLabelsEnabled(True)

# =============================================================================
# GRUPOS DE HOSTILIDADE
# Cada grupo agrega as hostilidades equivalentes para fins de camada
# =============================================================================

GRUPOS_HOSTILIDADE = {
    'COP — Amigo':        ['AMIGO', 'PRESUMIDO'],
    'COP — Hostil':       ['HOSTIL', 'SUSPEITO'],
    'COP — Neutro':       ['NEUTRO'],
    'COP — Desconhecido': ['DESCONHECIDO'],
}

# Ordem de empilhamento no painel (primeira = mais abaixo)
ORDEM_GRUPOS = ['COP — Desconhecido', 'COP — Neutro', 'COP — Amigo', 'COP — Hostil']


def _remover_grupo_cop(raiz):
    """Remove o grupo COP e todas as camadas filhas se existir."""
    grupo = raiz.findGroup('COP Tático')
    if grupo:
        grupo.removeAllChildren()
        raiz.removeChildNode(grupo)


def _criar_camada_vazia(nome):
    """Cria camada de pontos em memória com todos os campos."""
    campos = _definir_campos()
    camada = QgsVectorLayer(f'Point?crs={CRS_ENTRADA}', nome, 'memory')
    prov = camada.dataProvider()
    prov.addAttributes(campos)
    camada.updateFields()
    return camada


# =============================================================================
# CONFIGURAÇÃO DAS CAMADAS POR HOSTILIDADE
# Ordem de renderização: Hostil no topo, Amigo na base
# =============================================================================

GRUPOS_HOSTILIDADE = {
    'COP — Hostil':       ['HOSTIL', 'SUSPEITO'],
    'COP — Desconhecido': ['DESCONHECIDO'],
    'COP — Neutro':       ['NEUTRO'],
    'COP — Amigo':        ['AMIGO', 'PRESUMIDO'],
}

# Ordem de inserção no grupo — o último inserido fica no topo do painel
# Queremos Hostil no topo, então inserimos Amigo primeiro
ORDEM_GRUPOS = [
    'COP — Amigo',
    'COP — Neutro',
    'COP — Desconhecido',
    'COP — Hostil',
]


def _criar_camada_vazia(nome_camada):
    """Cria uma camada de pontos em memória com os campos padrão."""
    camada = QgsVectorLayer(
        f'Point?crs={CRS_ENTRADA}', nome_camada, 'memory'
    )
    prov = camada.dataProvider()
    prov.addAttributes(_definir_campos())
    camada.updateFields()
    return camada


def _salvar_no_gpkg(camada_mem, nome_camada, caminho_gpkg):
    """
    Salva/substitui uma camada no GeoPackage e retorna a camada
    persistente carregada do arquivo.

    O GeoPackage é criado na primeira vez e atualizado nas seguintes.
    Cada camada ocupa uma tabela separada dentro do mesmo arquivo.
    """
    import os
    from qgis.core import QgsVectorFileWriter, QgsCoordinateTransformContext

    opcoes = QgsVectorFileWriter.SaveVectorOptions()
    opcoes.driverName    = 'GPKG'
    opcoes.fileEncoding  = 'UTF-8'
    opcoes.layerName     = nome_camada

    # Se o arquivo já existe, atualiza a camada (não sobrescreve o arquivo inteiro)
    if os.path.exists(caminho_gpkg):
        opcoes.actionOnExistingFile = QgsVectorFileWriter.CreateOrOverwriteLayer
    else:
        opcoes.actionOnExistingFile = QgsVectorFileWriter.CreateOrOverwriteFile

    resultado, erro, _, _ = QgsVectorFileWriter.writeAsVectorFormatV3(
        camada_mem,
        caminho_gpkg,
        QgsCoordinateTransformContext(),
        opcoes,
    )

    if resultado != QgsVectorFileWriter.NoError:
        print(f'  ⚠ Erro ao salvar "{nome_camada}" no GeoPackage: {erro}')
        return camada_mem  # fallback: retorna a camada em memória

    # Carrega a camada persistente do GeoPackage
    uri = f'{caminho_gpkg}|layername={nome_camada}'
    camada_gpkg = QgsVectorLayer(uri, nome_camada, 'ogr')
    if not camada_gpkg.isValid():
        print(f'  ⚠ Camada "{nome_camada}" inválida após salvar no GeoPackage')
        return camada_mem

    return camada_gpkg


def _remover_grupo_cop(raiz):
    """Remove o grupo COP Tático e todas as suas camadas do projeto."""
    grupo = raiz.findGroup('COP Tático')
    if not grupo:
        return
    # Remove todas as camadas dentro do grupo do projeto
    for node in grupo.findLayers():
        QgsProject.instance().removeMapLayer(node.layerId())
    # Remove o grupo vazio
    raiz.removeChildNode(grupo)


# =============================================================================
# FUNÇÃO PRINCIPAL
# =============================================================================

def importar_cop(nome=NOME_CAMADA):
    """
    Lê o Google Sheets (ou fallback local) e cria camadas separadas
    por grupo de hostilidade, organizadas num grupo 'COP Tático' no
    painel de camadas do QGIS.
    """
    print(f'↺ Atualizando COP...')
    try:
        registros, fonte = ler_planilha_online()
    except Exception as e:
        print(f'❌ {e}')
        return None

    if not registros:
        print('⚠ Nenhum dado encontrado na planilha.')
        return None

    raiz = QgsProject.instance().layerTreeRoot()

    # Remove grupo anterior se existir
    _remover_grupo_cop(raiz)

    # Cria grupo novo no topo do painel
    grupo_cop = raiz.insertGroup(0, 'COP Tático')

    # Distribui registros por grupo de hostilidade
    por_grupo = {g: [] for g in GRUPOS_HOSTILIDADE}
    sem_grupo = []
    for reg in registros:
        host = reg.get('hostilidade', '').strip().upper()
        alocado = False
        for nome_grupo, hostilidades in GRUPOS_HOSTILIDADE.items():
            if host in hostilidades:
                por_grupo[nome_grupo].append(reg)
                alocado = True
                break
        if not alocado:
            sem_grupo.append(reg)

    if sem_grupo:
        print(f'  ⚠ {len(sem_grupo)} registros com hostilidade não reconhecida — ignorados')

    # Cria uma camada por grupo na ordem definida
    camadas_criadas = []
    total_feicoes = 0

    for nome_grupo in ORDEM_GRUPOS:
        regs_grupo = por_grupo[nome_grupo]
        camada = _criar_camada_vazia(nome_grupo)
        prov = camada.dataProvider()

        if regs_grupo:
            feicoes = _registros_para_feicoes(regs_grupo, camada.fields())
            prov.addFeatures(feicoes)
            total_feicoes += len(feicoes)
        
        camada.updateExtents()

        # Salva no GeoPackage — elimina o aviso de camada temporária
        camada = _salvar_no_gpkg(camada, nome_grupo, CAMINHO_GPKG)

        _aplicar_simbologia(camada)
        _aplicar_etiquetas(camada)

        # Adiciona ao projeto SEM mostrar no painel raiz
        QgsProject.instance().addMapLayer(camada, False)

        # Insere no grupo — Hostil no topo (último inserido = primeiro visível)
        grupo_cop.addLayer(camada)
        camadas_criadas.append(camada)

        n = len(regs_grupo)
        status = '✓' if n > 0 else '○'
        print(f'  {status} {nome_grupo}: {n} elementos')

    print(f'✓ COP carregada: {total_feicoes} elementos [{fonte}]')
    return camadas_criadas

# =============================================================================
# PUBLICAÇÃO NO GITHUB
# =============================================================================

def _publicar_github():
    """
    Chama publicar_github.py se existir em C:\\exercicio\\.
    Importado dinamicamente — silencioso se não estiver configurado.
    """
    script = r'C:\exercicio\publicar_github.py'
    if not os.path.exists(script):
        return
    try:
        from pathlib import Path
        ns = {'__file__': script, 'iface': None}
        exec(compile(Path(script).read_text(encoding='utf-8'), script, 'exec'), ns)
    except Exception as e:
        print(f'  ⚠ Publicação GitHub: {e}')


# =============================================================================
# ATUALIZAÇÃO AUTOMÁTICA
# Polling baseado em hash do conteúdo — só recarrega se os dados mudaram
# (não depende de mtime como na v5, funciona com fonte remota)
# =============================================================================

class CopAtualizador:
    def __init__(self, intervalo=INTERVALO_ATUALIZACAO):
        self.intervalo    = intervalo * 1000
        self._ultimo_hash = ''
        self._timer       = QTimer()
        self._timer.timeout.connect(self._verificar_e_atualizar)

    def iniciar(self):
        importar_cop()
        self._ultimo_hash = self._hash_atual()
        self._timer.start(self.intervalo)
        print(f'▶ Polling ativo — intervalo: {self.intervalo//1000}s')
        print(f'  Fonte: Google Sheets ({SHEETS_ID[:20]}...)')

    def parar(self):
        self._timer.stop()
        print('■ Polling pausado.')

    def _hash_atual(self):
        """Baixa a planilha e retorna o hash MD5 do conteúdo."""
        dados = baixar_sheets(_montar_url_sheets())
        if dados:
            return hashlib.md5(dados).hexdigest()
        return self._ultimo_hash  # mantém hash anterior se offline

    def _verificar_e_atualizar(self):
        novo_hash = self._hash_atual()
        if novo_hash == self._ultimo_hash:
            return  # sem mudança
        self._ultimo_hash = novo_hash
        print('↺ Mudança detectada no Sheets — recarregando...')
        importar_cop()
        _publicar_github()


# =============================================================================
# PONTO DE ENTRADA
#
# MODO 1 — carga única:
#     importar_cop()
#
# MODO 2 — atualização automática a cada 30s:
#     atualizador = CopAtualizador()
#     atualizador.iniciar()
#     # Para parar: atualizador.parar()
#
# MODO 3 — intervalo diferente (ex: 60s):
#     atualizador = CopAtualizador(intervalo=60)
#     atualizador.iniciar()
# =============================================================================

importar_cop()
